import openpyxl

wb = openpyxl.load_workbook('Course Content template.xlsx')
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

sheet = wb['Sheet1']
sheet = wb.active

b2_cell = sheet['B2']
print(b2_cell.value)
print(sheet['B2'].parent)
print(sheet.max_row, sheet.max_column)
print(sheet.dimensions)

for a, b, c, d, e, f in sheet[sheet.dimensions]:
    print(a.value, b.value, c.value, d.value, e.value, f.value)

for row in sheet.rows:
    for cell in row:
        print(f'{cell.value}', end='')
    print("\n")

print(sheet['D3'].value)
sheet['D4'] = 'Completed'
wb.save('Course Content template.xlsx')
print(sheet['C4'].value)